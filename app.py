import streamlit as st
import pandas as pd
import json
import os
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── CONFIG ────────────────────────────────────────────────────────────
st.set_page_config(page_title="VAV Commissioning Tool", page_icon="🏗️", layout="wide")
DATA_DIR = "data"
os.makedirs(DATA_DIR, exist_ok=True)

JOB_CODES = ["RH", "BMS", "HVAC", "MEP", "TAB", "CX"]

COLUMNS = [
    "Room #", "VAV #", "Trunk", "MAC Address", "Device Instance",
    "Box Size", "Min CFM", "Max CFM", "Reheat CFM", "Fan Type",
    "Supply Temp Sensor", "Space Temp Sensor", "Space Temp Actual",
    "Damper OPN CFM", "Damper CLSD CFM",
    "Heat Stage 1 ON SAT", "Heat Stage 1 OFF SAT",
    "Heat Stage 2 ON SAT", "Heat Stage 2 OFF SAT",
    "HW Valve OPN SAT", "HW Valve CLS SAT",
    "Spot-check Wiring", "Correction Factor",
    "Checked By", "Date (DD/MM/YY)", "Notes / Comments"
]

NUMERIC_COLS = [
    "Min CFM", "Max CFM", "Reheat CFM", "Space Temp Actual",
    "Damper OPN CFM", "Damper CLSD CFM",
    "Heat Stage 1 ON SAT", "Heat Stage 1 OFF SAT",
    "Heat Stage 2 ON SAT", "Heat Stage 2 OFF SAT",
    "HW Valve OPN SAT", "HW Valve CLS SAT", "Correction Factor"
]


# ─── HELPER FUNCTIONS ──────────────────────────────────────────────────
def get_jobs_file():
    return os.path.join(DATA_DIR, "jobs_index.json")


def load_jobs_index():
    f = get_jobs_file()
    if os.path.exists(f):
        with open(f, "r") as fh:
            return json.load(fh)
    return {}


def save_jobs_index(index):
    with open(get_jobs_file(), "w") as fh:
        json.dump(index, fh, indent=2)


def load_job_data(job_id):
    f = os.path.join(DATA_DIR, f"{job_id}.csv")
    if os.path.exists(f):
        df = pd.read_csv(f, dtype=str).fillna("")
        # Ensure all columns exist
        for c in COLUMNS:
            if c not in df.columns:
                df[c] = ""
        return df[COLUMNS]
    return pd.DataFrame(columns=COLUMNS)


def save_job_data(job_id, df):
    f = os.path.join(DATA_DIR, f"{job_id}.csv")
    df.to_csv(f, index=False)


def delete_job(job_id):
    idx = load_jobs_index()
    if job_id in idx:
        del idx[job_id]
        save_jobs_index(idx)
    f = os.path.join(DATA_DIR, f"{job_id}.csv")
    if os.path.exists(f):
        os.remove(f)


def generate_excel_report(job_name, job_code, df):
    """Generate a professional Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "VAV Commissioning"

    # ── Styles ──
    dark_blue = "1F4E79"
    med_blue = "2E75B6"
    light_blue = "D6E4F0"
    white = "FFFFFF"
    light_gray = "F2F7FB"
    green_bg = "C6EFCE"
    green_font = "006100"
    red_bg = "FFC7CE"
    red_font = "9C0006"

    title_font = Font(name="Calibri", size=18, bold=True, color=dark_blue)
    subtitle_font = Font(name="Calibri", size=11, bold=True, color=med_blue)
    header_font = Font(name="Calibri", size=9, bold=True, color=white)
    header_fill = PatternFill(start_color=dark_blue, end_color=dark_blue, fill_type="solid")
    cell_font = Font(name="Calibri", size=9)
    alt_fill = PatternFill(start_color=light_gray, end_color=light_gray, fill_type="solid")
    pass_fill = PatternFill(start_color=green_bg, end_color=green_bg, fill_type="solid")
    pass_font = Font(name="Calibri", size=9, color=green_font, bold=True)
    fail_fill = PatternFill(start_color=red_bg, end_color=red_bg, fill_type="solid")
    fail_font = Font(name="Calibri", size=9, color=red_font, bold=True)
    info_label_font = Font(name="Calibri", size=10, bold=True, color=dark_blue)
    info_value_font = Font(name="Calibri", size=10)
    footer_font = Font(name="Calibri", size=8, italic=True, color="808080")

    thin_border = Border(
        left=Side(style="thin", color=light_blue),
        right=Side(style="thin", color=light_blue),
        top=Side(style="thin", color=light_blue),
        bottom=Side(style="thin", color=light_blue),
    )
    header_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    num_cols = len(COLUMNS)
    now = datetime.now().strftime("%d/%m/%y")

    # ── Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value="VAV COMMISSIONING REPORT")
    c.font = title_font
    c.alignment = left_align
    ws.row_dimensions[1].height = 40

    # ── Subtitle ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c = ws.cell(row=2, column=1, value="Building Automation System — Functional Performance Test")
    c.font = subtitle_font
    c.alignment = left_align
    ws.row_dimensions[2].height = 24

    # ── Blank ──
    ws.row_dimensions[3].height = 8

    # ── Job Info ──
    ws.cell(row=4, column=1, value="Job Name:").font = info_label_font
    ws.cell(row=4, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=4, start_column=2, end_row=4, end_column=5)
    ws.cell(row=4, column=2, value=job_name).font = info_value_font
    info_border = Border(bottom=Side(style="thin", color=med_blue))
    for col in range(2, 6):
        ws.cell(row=4, column=col).border = info_border

    ws.cell(row=4, column=num_cols - 2, value="Report Date:").font = info_label_font
    ws.cell(row=4, column=num_cols - 2).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=4, start_column=num_cols - 1, end_row=4, end_column=num_cols)
    ws.cell(row=4, column=num_cols - 1, value=now).font = info_value_font
    for col in range(num_cols - 1, num_cols + 1):
        ws.cell(row=4, column=col).border = info_border

    ws.cell(row=5, column=1, value="Job Code:").font = info_label_font
    ws.cell(row=5, column=1).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=5)
    ws.cell(row=5, column=2, value=job_code).font = info_value_font
    for col in range(2, 6):
        ws.cell(row=5, column=col).border = info_border

    ws.cell(row=5, column=num_cols - 2, value="Total Units:").font = info_label_font
    ws.cell(row=5, column=num_cols - 2).alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(start_row=5, start_column=num_cols - 1, end_row=5, end_column=num_cols)
    ws.cell(row=5, column=num_cols - 1, value=str(len(df))).font = info_value_font
    for col in range(num_cols - 1, num_cols + 1):
        ws.cell(row=5, column=col).border = info_border

    # ── Blank ──
    ws.row_dimensions[6].height = 8

    # ── Headers ──
    header_row = 7
    ws.row_dimensions[header_row].height = 36
    for i, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(row=header_row, column=i, value=col_name)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center_align
        c.border = header_border

    # ── Data Rows ──
    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = header_row + 1 + row_idx
        ws.row_dimensions[excel_row].height = 22
        is_alt = row_idx % 2 == 1
        for col_idx, col_name in enumerate(COLUMNS, 1):
            val = str(row.get(col_name, ""))
            c = ws.cell(row=excel_row, column=col_idx, value=val)
            c.font = cell_font
            c.alignment = center_align
            c.border = thin_border
            if is_alt:
                c.fill = alt_fill
            # Color code spot-check wiring
            if col_name == "Spot-check Wiring":
                if val == "OK":
                    c.fill = pass_fill
                    c.font = pass_font
                elif val == "FAIL":
                    c.fill = fail_fill
                    c.font = fail_font

    # ── Footer ──
    footer_row = header_row + len(df) + 2
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=num_cols)
    c = ws.cell(row=footer_row, column=1, value=f"Generated by VAV Commissioning Tool | {now} | Confidential")
    c.font = footer_font
    c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Summary Section ──
    summary_row = footer_row + 2
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=6)
    c = ws.cell(row=summary_row, column=1, value="COMMISSIONING SUMMARY")
    c.font = Font(name="Calibri", size=12, bold=True, color=dark_blue)

    wiring_col = df.get("Spot-check Wiring", pd.Series(dtype=str))
    ok_count = (wiring_col == "OK").sum()
    fail_count = (wiring_col == "FAIL").sum()
    na_count = (wiring_col == "N/A").sum()
    pending = len(df) - ok_count - fail_count - na_count

    summary_items = [
        ("Total VAV Units:", str(len(df))),
        ("Wiring OK:", str(ok_count)),
        ("Wiring FAIL:", str(fail_count)),
        ("Wiring N/A:", str(na_count)),
        ("Pending:", str(pending)),
    ]
    for i, (label, value) in enumerate(summary_items):
        r = summary_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = info_label_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=r, column=2, value=value).font = info_value_font

    # ── Column Widths ──
    col_widths = {
        "Room #": 10, "VAV #": 10, "Trunk": 8, "MAC Address": 18,
        "Device Instance": 14, "Box Size": 10, "Min CFM": 10, "Max CFM": 10,
        "Reheat CFM": 12, "Fan Type": 10, "Supply Temp Sensor": 16,
        "Space Temp Sensor": 16, "Space Temp Actual": 14, "Damper OPN CFM": 14,
        "Damper CLSD CFM": 14, "Heat Stage 1 ON SAT": 16, "Heat Stage 1 OFF SAT": 16,
        "Heat Stage 2 ON SAT": 16, "Heat Stage 2 OFF SAT": 16,
        "HW Valve OPN SAT": 16, "HW Valve CLS SAT": 14,
        "Spot-check Wiring": 16, "Correction Factor": 14,
        "Checked By": 12, "Date (DD/MM/YY)": 14, "Notes / Comments": 24,
    }
    for i, col_name in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col_name, 12)

    # ── Print Setup ──
    ws.sheet_properties.pageSetUpPr = None
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── SESSION STATE INIT ───────────────────────────────────────────────
if "current_job" not in st.session_state:
    st.session_state.current_job = None
if "page" not in st.session_state:
    st.session_state.page = "home"


# ─── CUSTOM CSS ────────────────────────────────────────────────────────
st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; max-width: 100%; }
    div[data-testid="stDataEditor"] { border: 1px solid #334155; border-radius: 8px; }
    .stButton button { border-radius: 8px; font-weight: 500; }
    h1 { color: #1e40af; }
    .job-card {
        background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 10px;
        padding: 16px; margin-bottom: 8px; transition: all 0.2s;
    }
    .job-card:hover { border-color: #3b82f6; box-shadow: 0 2px 8px rgba(59,130,246,0.1); }
    .metric-card {
        background: linear-gradient(135deg, #1e40af 0%, #3b82f6 100%);
        border-radius: 10px; padding: 20px; color: white; text-align: center;
    }
    .metric-card h3 { margin: 0; font-size: 28px; }
    .metric-card p { margin: 4px 0 0 0; font-size: 12px; opacity: 0.8; }
</style>
""", unsafe_allow_html=True)


# ─── HOME PAGE ─────────────────────────────────────────────────────────
def show_home():
    st.markdown("# 🏗️ VAV Commissioning Tool")
    st.markdown("**Building Automation System — Centralized Commissioning Platform**")
    st.divider()

    jobs = load_jobs_index()

    # ── Metrics ──
    total_jobs = len(jobs)
    total_units = 0
    for jid in jobs:
        df = load_job_data(jid)
        total_units += len(df)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown(f"""<div class="metric-card"><h3>{total_jobs}</h3><p>TOTAL JOBS</p></div>""", unsafe_allow_html=True)
    with col2:
        st.markdown(f"""<div class="metric-card"><h3>{total_units}</h3><p>TOTAL VAV UNITS</p></div>""", unsafe_allow_html=True)
    with col3:
        st.markdown(f"""<div class="metric-card"><h3>{len(JOB_CODES)}</h3><p>JOB CODES</p></div>""", unsafe_allow_html=True)

    st.markdown("")

    # ── Create New Job ──
    st.subheader("➕ Create New Job")
    with st.container(border=True):
        c1, c2, c3 = st.columns([3, 1, 1])
        with c1:
            new_name = st.text_input("Job Name", placeholder="e.g. Tower A Level 12")
        with c2:
            new_code = st.selectbox("Job Code", JOB_CODES)
        with c3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🚀 Create Job", use_container_width=True, type="primary"):
                if new_name.strip():
                    jid = f"job_{int(datetime.now().timestamp()*1000)}"
                    idx = load_jobs_index()
                    idx[jid] = {
                        "name": new_name.strip(),
                        "code": new_code,
                        "created": datetime.now().isoformat()
                    }
                    save_jobs_index(idx)
                    save_job_data(jid, pd.DataFrame(columns=COLUMNS))
                    st.session_state.current_job = jid
                    st.session_state.page = "form"
                    st.rerun()
                else:
                    st.error("Please enter a job name.")

    # ── Existing Jobs ──
    st.subheader(f"📋 Saved Jobs ({total_jobs})")

    if not jobs:
        st.info("No jobs yet. Create one above to get started!")
    else:
        search = st.text_input("🔍 Search jobs", placeholder="Filter by name or code...")
        for jid, jinfo in sorted(jobs.items(), key=lambda x: x[1].get("created", ""), reverse=True):
            name = jinfo.get("name", "Untitled")
            code = jinfo.get("code", "—")
            created = jinfo.get("created", "")
            if search and search.lower() not in name.lower() and search.lower() not in code.lower():
                continue
            df = load_job_data(jid)
            created_str = ""
            if created:
                try:
                    created_str = datetime.fromisoformat(created).strftime("%d %b %Y")
                except:
                    pass
            with st.container(border=True):
                c1, c2, c3 = st.columns([4, 1, 1])
                with c1:
                    st.markdown(f"**`{code}`** — {name}")
                    st.caption(f"{len(df)} units · Created {created_str}")
                with c2:
                    if st.button("📂 Open", key=f"open_{jid}", use_container_width=True):
                        st.session_state.current_job = jid
                        st.session_state.page = "form"
                        st.rerun()
                with c3:
                    if st.button("🗑️ Delete", key=f"del_{jid}", use_container_width=True):
                        delete_job(jid)
                        st.rerun()


# ─── FORM PAGE ─────────────────────────────────────────────────────────
def show_form():
    jid = st.session_state.current_job
    jobs = load_jobs_index()
    if jid not in jobs:
        st.session_state.page = "home"
        st.rerun()
        return

    jinfo = jobs[jid]
    job_name = jinfo["name"]
    job_code = jinfo["code"]

    # ── Top Bar ──
    c1, c2, c3, c4 = st.columns([1, 3, 1, 1])
    with c1:
        if st.button("← Back to Jobs"):
            st.session_state.page = "home"
            st.rerun()
    with c2:
        st.markdown(f"### `{job_code}` — {job_name}")
    with c3:
        add_rows = st.number_input("Add rows", min_value=1, max_value=50, value=1, label_visibility="collapsed")
    with c4:
        add_btn = st.button(f"➕ Add {add_rows} Row(s)", use_container_width=True)

    st.divider()

    # ── Load Data ──
    df = load_job_data(jid)

    if add_btn:
        new_rows = pd.DataFrame([{c: "" for c in COLUMNS}] * add_rows)
        df = pd.concat([df, new_rows], ignore_index=True)
        save_job_data(jid, df)
        st.rerun()

    # ── Column Config for Data Editor ──
    column_config = {
        "Spot-check Wiring": st.column_config.SelectboxColumn(
            "Spot-check Wiring",
            options=["OK", "FAIL", "N/A"],
            width="medium"
        ),
        "Fan Type": st.column_config.SelectboxColumn(
            "Fan Type",
            options=["Series", "Parallel", "None", "ECM", "PSC"],
            width="small"
        ),
        "Date (DD/MM/YY)": st.column_config.TextColumn(
            "Date (DD/MM/YY)",
            width="medium"
        ),
        "Notes / Comments": st.column_config.TextColumn(
            "Notes / Comments",
            width="large"
        ),
    }

    # Make numeric columns use number input
    for nc in NUMERIC_COLS:
        column_config[nc] = st.column_config.NumberColumn(nc, width="small")

    # ── Data Editor ──
    edited_df = st.data_editor(
        df,
        column_config=column_config,
        num_rows="dynamic",
        use_container_width=True,
        height=500,
        key=f"editor_{jid}",
    )

    # ── Action Buttons ──
    st.markdown("")
    c1, c2, c3, c4 = st.columns(4)

    with c1:
        if st.button("💾 Save Data", use_container_width=True, type="primary"):
            save_job_data(jid, edited_df)
            st.success("✅ Data saved!")

    with c2:
        excel_buf = generate_excel_report(job_name, job_code, edited_df)
        filename = f"VAV_Commissioning_{job_code}_{job_name.replace(' ', '_')}.xlsx"
        st.download_button(
            "📊 Export Excel Report",
            data=excel_buf,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
            use_container_width=True,
        )

    with c3:
        csv_data = edited_df.to_csv(index=False).encode("utf-8")
        st.download_button(
            "📄 Export CSV",
            data=csv_data,
            file_name=f"VAV_{job_code}_{job_name.replace(' ', '_')}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with c4:
        if st.button("🔄 Refresh", use_container_width=True):
            st.rerun()

    # ── Summary Stats ──
    st.divider()
    st.subheader("📊 Quick Summary")
    wiring = edited_df.get("Spot-check Wiring", pd.Series(dtype=str))
    ok = (wiring == "OK").sum()
    fail = (wiring == "FAIL").sum()
    na = (wiring == "N/A").sum()
    pending = len(edited_df) - ok - fail - na

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("Total Units", len(edited_df))
    m2.metric("Wiring OK ✅", ok)
    m3.metric("Wiring FAIL ❌", fail)
    m4.metric("Wiring N/A", na)
    m5.metric("Pending ⏳", pending)

    # Progress bar
    if len(edited_df) > 0:
        progress = (ok + fail + na) / len(edited_df)
        st.progress(progress, text=f"Commissioning Progress: {progress*100:.0f}%")


# ─── MAIN ROUTER ──────────────────────────────────────────────────────
if st.session_state.page == "form" and st.session_state.current_job:
    show_form()
else:
    show_home()
